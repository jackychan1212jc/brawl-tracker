import urllib.parse
import requests
import time
import threading
from collections import deque
import os
import re
import webbrowser
import json
import socket
import shutil
from datetime import datetime, timedelta
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

# --- 新增的 Web 伺服器套件 ---
from fastapi import FastAPI
from fastapi.responses import FileResponse, HTMLResponse
import uvicorn

# --- 1. 基本設定 ---
API_TOKEN = 'eyJ0eXAiOiJKV1QiLCJhbGciOiJIUzUxMiIsImtpZCI6IjI4YTMxOGY3LTAwMDAtYTFlYi03ZmExLTJjNzQzM2M2Y2NhNSJ9.eyJpc3MiOiJzdXBlcmNlbGwiLCJhdWQiOiJzdXBlcmNlbGw6Z2FtZWFwaSIsImp0aSI6IjZlZDFjMzQ0LTFlMDEtNGIwNS04NzNkLTEzODNiZDI5ZjBlNSIsImlhdCI6MTc4Njk4ODU3Niwic3ViIjoiZGV2ZWxvcGVyLzA4YmE4NjhhLTNlOGItNDMwMS1iNmE2LWJkODExMTgxZTliNyIsInNjb3BlcyI6WyJicmF3bHN0YXJzIl0sImxpbWl0cyI6W3sidGllciI6ImRldmVsb3Blci9zaWx2ZXIiLCJ0eXBlIjoidGhyb3R0bGluZyJ9LHsiY2lkcnMiOlsiNjAuMjQ2LjE3NC4xMjUiXSwidHlwZSI6ImNsaWVudCJ9XX0.FG5pYWvXFqxLEF20SyIgjzGdDrAuod6C36gTJ6isVoQQ_-7SrDboWO-Y6qMc_pOH0mdsM46KdUX6p11mp64W0w'

ACCOUNTS = {
    '大號': {'tag': '#9P2GP0UL9', 'excel': r"D:\我的荒野亂鬥戰績_自動更新版.xlsx"},
    '小號': {'tag': '#2QGP2L0VP', 'excel': r"D:\我的荒野亂鬥戰績_小號_自動更新版.xlsx"}
}

headers_official = {'Authorization': f'Bearer {API_TOKEN}'}
headers_ninja = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
}

MODE_TRANSLATION = {
    'gemGrab': '寶石爭奪戰', 'brawlBall': '亂鬥足球', 'bounty': '搶星大作戰',
    'heist': '金庫攻防戰', 'hotZone': '據點搶奪戰', 'knockout': '極限淘汰賽',
    'wipeout': '積分爭奪戰', 'duels': '亂鬥擂台', 'soloShowdown': '單人生死鬥',
    'duoShowdown': '雙人生死鬥', 'basketBrawl': '亂鬥籃球', 'payload': '礦車競速',
    'lastStand': 'megaBoss', 'bossFight': '團隊首領戰', 'roboRumble': '機甲入侵', 
    'bigGame': '巨型獵場', 'brawlArena': '亂鬥競技場', 'arena': '亂鬥競技場'
}

PVE_MODES = ['lastStand', 'bossFight', 'roboRumble', 'bigGame', 'megaBoss']
TARGET_SIX_MODES = ['搶星大作戰', '寶石爭奪戰', '金庫攻防戰', '亂鬥足球', '據點搶奪戰', '極限淘汰賽']

startup_time_local = datetime.utcnow() + timedelta(hours=8)
current_local_time_str = startup_time_local.strftime('%Y-%m-%d %H:%M:%S')

account_stats = {
    '大號': {'start_trophies': None, 'start_elo': None, 'last_time': None, 'last_raw_time': "", 'current_trophies': 0, 'diff_str': "+0", 'elo_str': "尚未更新", 'elo_diff_str': "+0", '3v3_victories': 0, 'elo_tier': 'UNKNOWN', 'ui_session': {}, 'ui_all_time': {}, 'ranked_seasons_session': {}, 'ranked_seasons_all_time': {}, 'startup_formatted_time': current_local_time_str, 'owned_brawlers': []},
    '小號': {'start_trophies': None, 'start_elo': None, 'last_time': None, 'last_raw_time': "", 'current_trophies': 0, 'diff_str': "+0", 'elo_str': "尚未更新", 'elo_diff_str': "+0", '3v3_victories': 0, 'elo_tier': 'UNKNOWN', 'ui_session': {}, 'ui_all_time': {}, 'ranked_seasons_session': {}, 'ranked_seasons_all_time': {}, 'startup_formatted_time': current_local_time_str, 'owned_brawlers': []}
}

# 補回網頁版生成所需的預設變數
current_display_account = '大號'
current_view_mode = 'session'

base_dir = r"D:\Brawl Stars Win Rate\py\Brawl_Tactics_Dashboard"
if not os.path.exists(base_dir):
    try: os.makedirs(base_dir)
    except: pass

last_sync_time_str = "等待同步..."
last_new_data_real_time = time.time()

network_ping_api = 0
network_ping_tokyo = 0
network_ping_hk = 0
MAX_PING_HISTORY = 60
ping_history = {
    'api': deque(maxlen=MAX_PING_HISTORY),
    'tokyo': deque(maxlen=MAX_PING_HISTORY),
    'hk': deque(maxlen=MAX_PING_HISTORY)
}

def get_wr(w, l, d=0):
    total = w + l + d
    return f"{w/total*100:.1f}%" if total > 0 else "0.0%"

def group_ranked_sets(df):
    if df.empty: return df
    df = df.copy()
    df['對戰時間_dt'] = pd.to_datetime(df['對戰時間'], errors='coerce')
    df = df.sort_values('對戰時間_dt', ascending=True).reset_index(drop=True)
    processed_rows = []
    current_set = []
    
    def flush_set(curr_set):
        if not curr_set: return []
        v_count = sum(1 for row in curr_set if row['戰果'] == 'victory')
        d_count = sum(1 for row in curr_set if row['戰果'] == 'defeat')
        if v_count >= 2 or d_count >= 2:
            rep_row = curr_set[-1].copy()
            rep_row['戰果'] = 'victory' if v_count >= 2 else 'defeat'
            return [rep_row]
        else: return []

    for idx, row in df.iterrows():
        is_ranked = row.get('類型') in ['soloRanked', 'teamRanked']
        mech = row.get('排位機制', 'BO1')
        if not is_ranked or mech != 'BO3':
            if current_set:
                processed_rows.extend(flush_set(current_set))
                current_set = []
            processed_rows.append(row)
            continue
        if not current_set: current_set.append(row)
        else:
            last_row = current_set[-1]
            same_type = row['類型'] == last_row['類型']
            same_map = row['地圖'] == last_row['地圖']
            same_brawler = row['我方英雄'] == last_row['我方英雄']
            time_diff = abs((row['對戰時間_dt'] - last_row['對戰時間_dt']).total_seconds()) / 60.0
            if same_type and same_map and same_brawler and time_diff <= 15:
                current_set.append(row)
                v_c = sum(1 for r in current_set if r['戰果'] == 'victory')
                d_c = sum(1 for r in current_set if r['戰果'] == 'defeat')
                if v_c >= 2 or d_c >= 2:
                    processed_rows.extend(flush_set(current_set))
                    current_set = []
            else:
                processed_rows.extend(flush_set(current_set))
                current_set = [row]
                
    if current_set: processed_rows.extend(flush_set(current_set))
    df_grouped = pd.DataFrame(processed_rows)
    if not df_grouped.empty:
        df_grouped = df_grouped.drop(columns=['對戰時間_dt'])
        df_grouped = df_grouped.sort_values('對戰時間', ascending=False)
    return df_grouped

def process_and_group_dataframe(df):
    if df.empty: return df
    df_valid = df[(df['戰果'].isin(['victory', 'defeat', 'draw'])) & (df['類型'] != 'friendly')].copy()
    df_grouped = group_ranked_sets(df_valid)
    if df_grouped.empty: return df_grouped
    
    def determine_classification(row):
        raw_type = str(row.get('類型', ''))
        raw_mode = str(row.get('模式', ''))
        is_box_event = row.get('是否寶箱活動', False)
        my_team = str(row.get('我方陣容', ''))
        enemy_team = str(row.get('敵方陣容', ''))
        brawlers = [x.strip() for x in (my_team + ',' + enemy_team).split(',') if x.strip() and x.strip() != 'nan']
        is_mirror = len(brawlers) > 1 and len(set(brawlers)) == 1
        base_mode_zh = MODE_TRANSLATION.get(raw_mode, raw_mode)
        
        if is_mirror: ui_type = '鏡像亂鬥'
        elif is_box_event: ui_type = '寶箱活動'
        elif raw_type == 'challenge': ui_type = '挑戰'
        elif raw_mode in PVE_MODES: ui_type = '特別活動'
        elif raw_type in ['soloRanked', 'teamRanked']: ui_type = '排位賽'
        elif raw_type == 'ranked': ui_type = '一般模式'
        else: ui_type = '特別活動'
        return pd.Series([ui_type, base_mode_zh])
        
    df_grouped[['UI動態分類', '模式中文']] = df_grouped.apply(determine_classification, axis=1)
    return df_grouped

def build_ui_dict(df_grouped):
    ui_data = {'r_wins': 0, 'r_losses': 0, 'r_draws': 0, 't_wins': 0, 't_losses': 0, 't_draws': 0,
               'box_wins': 0, 'box_losses': 0, 'box_draws': 0, 'c_wins': 0, 'c_losses': 0, 'c_draws': 0,
               's_wins': 0, 's_losses': 0, 's_draws': 0, 'm_wins': 0, 'm_losses': 0, 'm_draws': 0, 'brawler_stats': {}, 'map_stats': {}}
    if df_grouped is None or df_grouped.empty: return ui_data

    ui_data['r_wins'] = len(df_grouped[(df_grouped['UI動態分類'] == '排位賽') & (df_grouped['戰果'] == 'victory')])
    ui_data['r_losses'] = len(df_grouped[(df_grouped['UI動態分類'] == '排位賽') & (df_grouped['戰果'] == 'defeat')])
    ui_data['r_draws'] = len(df_grouped[(df_grouped['UI動態分類'] == '排位賽') & (df_grouped['戰果'] == 'draw')])
    ui_data['t_wins'] = len(df_grouped[(df_grouped['UI動態分類'] == '一般模式') & (df_grouped['戰果'] == 'victory')])
    ui_data['t_losses'] = len(df_grouped[(df_grouped['UI動態分類'] == '一般模式') & (df_grouped['戰果'] == 'defeat')])
    ui_data['t_draws'] = len(df_grouped[(df_grouped['UI動態分類'] == '一般模式') & (df_grouped['戰果'] == 'draw')])
    ui_data['box_wins'] = len(df_grouped[(df_grouped['UI動態分類'] == '寶箱活動') & (df_grouped['戰果'] == 'victory')])
    ui_data['box_losses'] = len(df_grouped[(df_grouped['UI動態分類'] == '寶箱活動') & (df_grouped['戰果'] == 'defeat')])
    ui_data['box_draws'] = len(df_grouped[(df_grouped['UI動態分類'] == '寶箱活動') & (df_grouped['戰果'] == 'draw')])
    ui_data['c_wins'] = len(df_grouped[(df_grouped['UI動態分類'] == '挑戰') & (df_grouped['戰果'] == 'victory')])
    ui_data['c_losses'] = len(df_grouped[(df_grouped['UI動態分類'] == '挑戰') & (df_grouped['戰果'] == 'defeat')])
    ui_data['c_draws'] = len(df_grouped[(df_grouped['UI動態分類'] == '挑戰') & (df_grouped['戰果'] == 'draw')])
    ui_data['s_wins'] = len(df_grouped[(df_grouped['UI動態分類'] == '特別活動') & (df_grouped['戰果'] == 'victory')])
    ui_data['s_losses'] = len(df_grouped[(df_grouped['UI動態分類'] == '特別活動') & (df_grouped['戰果'] == 'defeat')])
    ui_data['s_draws'] = len(df_grouped[(df_grouped['UI動態分類'] == '特別活動') & (df_grouped['戰果'] == 'draw')])
    ui_data['m_wins'] = len(df_grouped[(df_grouped['UI動態分類'] == '鏡像亂鬥') & (df_grouped['戰果'] == 'victory')])
    ui_data['m_losses'] = len(df_grouped[(df_grouped['UI動態分類'] == '鏡像亂鬥') & (df_grouped['戰果'] == 'defeat')])
    ui_data['m_draws'] = len(df_grouped[(df_grouped['UI動態分類'] == '鏡像亂鬥') & (df_grouped['戰果'] == 'draw')])
    
    b_stats = {}
    m_stats = {}
    for _, row in df_grouped.iterrows():
        b = str(row['我方英雄']).upper()
        b_type = row['UI動態分類'] 
        b_mode = row['模式中文']
        res = row['戰果']
        
        if b not in ['NAN', 'NONE', '未知', '']:
            if b not in b_stats: b_stats[b] = {'W': 0, 'L': 0, 'D': 0, 'types': {}}
            if b_type not in b_stats[b]['types']: b_stats[b]['types'][b_type] = {'W': 0, 'L': 0, 'D': 0, 'modes': {}}
            if b_mode not in b_stats[b]['types'][b_type]['modes']: b_stats[b]['types'][b_type]['modes'][b_mode] = {'W': 0, 'L': 0, 'D': 0}
            if res == 'victory': 
                b_stats[b]['W'] += 1; b_stats[b]['types'][b_type]['W'] += 1; b_stats[b]['types'][b_type]['modes'][b_mode]['W'] += 1
            elif res == 'defeat': 
                b_stats[b]['L'] += 1; b_stats[b]['types'][b_type]['L'] += 1; b_stats[b]['types'][b_type]['modes'][b_mode]['L'] += 1
            elif res == 'draw': 
                b_stats[b]['D'] += 1; b_stats[b]['types'][b_type]['D'] += 1; b_stats[b]['types'][b_type]['modes'][b_mode]['D'] += 1
                
        if b_type not in m_stats: m_stats[b_type] = {'W': 0, 'L': 0, 'D': 0, 'modes': {}}
        if b_mode not in m_stats[b_type]['modes']: m_stats[b_type]['modes'][b_mode] = {'W': 0, 'L': 0, 'D': 0}
        if res == 'victory': m_stats[b_type]['W'] += 1; m_stats[b_type]['modes'][b_mode]['W'] += 1
        elif res == 'defeat': m_stats[b_type]['L'] += 1; m_stats[b_type]['modes'][b_mode]['L'] += 1
        elif res == 'draw': m_stats[b_type]['D'] += 1; m_stats[b_type]['modes'][b_mode]['D'] += 1
            
    ui_data['brawler_stats'] = b_stats
    ui_data['map_stats'] = m_stats
    return ui_data

def build_ranked_ui_dict(df_grouped):
    res = {}
    if df_grouped is None or df_grouped.empty: return res
    df_rk = df_grouped[df_grouped['UI動態分類'] == '排位賽'].copy()
    if df_rk.empty: return res
    
    if '賽季' not in df_rk.columns: df_rk['賽季'] = '未知賽季'
    
    for season, grp in df_rk.groupby('賽季'):
        s_season = str(season).strip()
        if s_season.endswith('.0'): s_season = s_season[:-2]
        if not s_season or s_season == 'nan': s_season = '未知賽季'
        
        start_date, end_date = "", ""
        if '對戰時間' in grp.columns:
            dates = grp['對戰時間'].dropna().astype(str).tolist()
            valid_dates = [d for d in dates if len(d) >= 10 and '-' in d[:10]]
            if valid_dates:
                valid_dates.sort()
                start_date = valid_dates[0][5:10].replace('-', '/')
                end_date = valid_dates[-1][5:10].replace('-', '/')
        
        s_w = len(grp[grp['戰果'] == 'victory'])
        s_l = len(grp[grp['戰果'] == 'defeat'])
        s_d = len(grp[grp['戰果'] == 'draw'])
        
        brawlers = {}
        for brawler, b_grp in grp.groupby('我方英雄'):
            b_w = len(b_grp[b_grp['戰果'] == 'victory'])
            b_l = len(b_grp[b_grp['戰果'] == 'defeat'])
            b_d = len(b_grp[b_grp['戰果'] == 'draw'])
            modes = {}
            for mode, m_grp in b_grp.groupby('模式中文'):
                m_w = len(m_grp[m_grp['戰果'] == 'victory'])
                m_l = len(m_grp[m_grp['戰果'] == 'defeat'])
                m_d = len(m_grp[m_grp['戰果'] == 'draw'])
                modes[mode] = {'w': m_w, 'l': m_l, 'd': m_d}
            brawlers[brawler] = {'w': b_w, 'l': b_l, 'd': b_d, 'modes': modes}
            
        res[s_season] = {
            'w': s_w, 'l': s_l, 'd': s_d, 
            'start_date': start_date, 'end_date': end_date, 
            'brawlers': brawlers
        }
    return res

def build_js_view_data(ui_data):
    r_wins, r_losses, r_draws = ui_data.get('r_wins', 0), ui_data.get('r_losses', 0), ui_data.get('r_draws', 0)
    t_wins, t_losses, t_draws = ui_data.get('t_wins', 0), ui_data.get('t_losses', 0), ui_data.get('t_draws', 0)
    box_w, box_l, box_d = ui_data.get('box_wins', 0), ui_data.get('box_losses', 0), ui_data.get('box_draws', 0)
    c_wins, c_losses, c_draws = ui_data.get('c_wins', 0), ui_data.get('c_losses', 0), ui_data.get('c_draws', 0)
    s_wins, s_losses, s_draws = ui_data.get('s_wins', 0), ui_data.get('s_losses', 0), ui_data.get('s_draws', 0)
    m_wins, m_losses, m_draws = ui_data.get('m_wins', 0), ui_data.get('m_losses', 0), ui_data.get('m_draws', 0)
    brawler_stats = ui_data.get('brawler_stats', {})
    
    total_wins = r_wins + t_wins + box_w + c_wins + s_wins + m_wins
    total_losses = r_losses + t_losses + box_l + c_losses + s_losses + m_losses
    total_draws = r_draws + t_draws + box_d + c_draws + s_draws + m_draws
    
    merged_s_wins = s_wins + box_w + c_wins + m_wins
    merged_s_losses = s_losses + box_l + c_losses + m_losses
    merged_s_draws = s_draws + box_d + c_draws + m_draws
    
    summary = {
        'ranked': {'txt': f"{r_wins}W - {r_losses}L ({get_wr(r_wins, r_losses, r_draws)})", 'w': r_wins, 'l': r_losses, 'd': r_draws},
        'casual': {'txt': f"{t_wins}W - {t_losses}L ({get_wr(t_wins, t_losses, t_draws)})", 'w': t_wins, 'l': t_losses, 'd': t_draws},
        'special': {'txt': f"{merged_s_wins}W - {merged_s_losses}L ({get_wr(merged_s_wins, merged_s_losses, merged_s_draws)})", 'w': merged_s_wins, 'l': merged_s_losses, 'd': merged_s_draws},
        'total': {'txt': f"{total_wins}W - {total_losses}L ({get_wr(total_wins, total_losses, total_draws)})", 'w': total_wins, 'l': total_losses, 'd': total_draws}
    }
    
    brawlers = []
    for b_type_zh, icon in [('排位賽', '🏅'), ('一般模式', '⏳'), ('挑戰', '🎯'), ('寶箱活動', '🎁'), ('特別活動', '🎪'), ('鏡像亂鬥', '🎭')]:
        type_brawlers = {}
        for b_name, b_data in brawler_stats.items():
            if b_type_zh in b_data.get('types', {}): type_brawlers[b_name] = b_data['types'][b_type_zh]
        if not type_brawlers: continue
        
        display_title = '全新英雄寶箱活動！' if b_type_zh == '寶箱活動' else b_type_zh
        cat_dict = {'icon': icon, 'title': display_title, 'items': []}
        
        sorted_brawlers = sorted(type_brawlers.items(), key=lambda item: (item[1]['W'] + item[1]['L'] + item[1]['D'], item[1]['W']), reverse=True)
        for b_name, b_stats_item in sorted_brawlers:
            w, l, d = b_stats_item['W'], b_stats_item['L'], b_stats_item['D']
            cat_dict['items'].append({'name': b_name.title(), 'stats': f"{w}W - {l}L ({get_wr(w, l, d)})", 'w': w, 'l': l, 'd': d})
        brawlers.append(cat_dict)
        
    brawler_details = {}
    for b_name, b_data in brawler_stats.items():
        tot_w, tot_l, tot_d = b_data['W'], b_data['L'], b_data['D']
        b_dict = {'summary': f"{tot_w}W - {tot_l}L ({get_wr(tot_w, tot_l, tot_d)})", 'w': tot_w, 'l': tot_l, 'd': tot_d, 'cats': []}
        for b_type_zh, icon in [('排位賽', '🏅'), ('一般模式', '⏳'), ('挑戰', '🎯'), ('寶箱活動', '🎁'), ('特別活動', '🎪'), ('鏡像亂鬥', '🎭')]:
            if b_type_zh in b_data.get('types', {}):
                t_data = b_data['types'][b_type_zh]
                cat_js = {'icon': icon, 'title': b_type_zh, 'wins': t_data['W'], 'losses': t_data['L'], 'wr': get_wr(t_data['W'], t_data['L'], t_data['D']), 'w': t_data['W'], 'l': t_data['L'], 'd': t_data['D'], 'modes': []}
                sorted_modes = sorted(t_data['modes'].items(), key=lambda x: TARGET_SIX_MODES.index(x[0]) if x[0] in TARGET_SIX_MODES else 99)
                for m_zh, m_d in sorted_modes:
                    cat_js['modes'].append({'name': m_zh, 'stats': f"{m_d['W']}W - {m_d['L']}L ({get_wr(m_d['W'], m_d['L'], m_d['D'])})", 'w': m_d['W'], 'l': m_d['L'], 'd': m_d['D']})
                b_dict['cats'].append(cat_js)
        brawler_details[b_name] = b_dict

    js_map_stats = []
    for icon, cat in [('🏅', '排位賽'), ('⏳', '一般模式')]:
        modes = {}
        if cat == '一般模式':
            for m in TARGET_SIX_MODES:
                mw = ui_data.get('map_stats', {}).get('一般模式', {}).get('modes', {}).get(m, {}).get('W', 0) + ui_data.get('map_stats', {}).get('寶箱活動', {}).get('modes', {}).get(m, {}).get('W', 0)
                ml = ui_data.get('map_stats', {}).get('一般模式', {}).get('modes', {}).get(m, {}).get('L', 0) + ui_data.get('map_stats', {}).get('寶箱活動', {}).get('modes', {}).get(m, {}).get('L', 0)
                md = ui_data.get('map_stats', {}).get('一般模式', {}).get('modes', {}).get(m, {}).get('D', 0) + ui_data.get('map_stats', {}).get('寶箱活動', {}).get('modes', {}).get(m, {}).get('D', 0)
                modes[m] = {'W': mw, 'L': ml, 'D': md}
        else:
            cat_data = ui_data.get('map_stats', {}).get(cat, {})
            for m in TARGET_SIX_MODES:
                if m in cat_data.get('modes', {}):
                    modes[m] = cat_data['modes'][m]
                else:
                    modes[m] = {'W': 0, 'L': 0, 'D': 0}
                    
        w = sum(v['W'] for v in modes.values())
        l = sum(v['L'] for v in modes.values())
        d = sum(v['D'] for v in modes.values())
                
        def get_wr_internal(w,l,d): return f"{w/(w+l+d)*100:.1f}%" if w+l+d>0 else "0.0%"
        cat_js = {'icon': icon, 'title': cat, 'wins': w, 'losses': l, 'wr': get_wr_internal(w, l, d), 'w': w, 'l': l, 'd': d, 'modes': []}
        
        for m in TARGET_SIX_MODES:
            md = modes[m]
            cat_js['modes'].append({'name': m, 'stats': f"{md['W']}W - {md['L']}L ({get_wr_internal(md['W'], md['L'], md['D'])})", 'w': md['W'], 'l': md['L'], 'd': md['D']})
        js_map_stats.append(cat_js)

    return {'summary': summary, 'brawlers': brawlers, 'brawler_details': brawler_details, 'map_stats': js_map_stats}

def generate_interactive_main_page(current_acc, current_view):
    app_data = {}
    for acc in ['大號', '小號']:
        stats = account_stats[acc]
        color = "#00FFAA" if acc == '大號' else "#00CCFF"
        acc_dict = {
            'color': color,
            'trophies': stats.get('current_trophies', 0),
            'diff_trophies': stats.get('diff_str', '+0'),
            'victories_3v3': stats.get('3v3_victories', 0),
            'elo': stats.get('elo_str', '尚未更新'),
            'diff_elo': stats.get('elo_diff_str', '+0'),
            'tier': stats.get('elo_tier', 'UNKNOWN'),
            'session': build_js_view_data(stats.get('ui_session', {})),
            'all_time': build_js_view_data(stats.get('ui_all_time', {})),
            'ranked_seasons_session': stats.get('ranked_seasons_session', {}),
            'ranked_seasons_all_time': stats.get('ranked_seasons_all_time', {})
        }
        app_data[acc] = acc_dict

    js_string = json.dumps(app_data, ensure_ascii=False)
    
    ping_payload = {
        'current': {'api': network_ping_api, 'tokyo': network_ping_tokyo, 'hk': network_ping_hk},
        'history': {
            'api': list(ping_history['api']),
            'tokyo': list(ping_history['tokyo']),
            'hk': list(ping_history['hk'])
        }
    }
    js_ping_string = json.dumps(ping_payload, ensure_ascii=False)
    
    html_template = """
    <!DOCTYPE html>
    <html lang="zh-TW">
    <head>
        <meta charset="UTF-8">
        <title>Brawl Tactics Dashboard</title>
        <style>
            body { background-color: #121212; color: #FFFFFF; font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; padding: 40px 8vw; margin: 0; display: flex; justify-content: center; }
            body.no-scroll { overflow: hidden; }

            .container { width: 100%; max-width: 900px; background-color: #1A1F24; border-radius: 15px; border: 1px solid #2A323C; padding: 40px; box-shadow: 0 10px 30px rgba(0,0,0,0.5); }
            
            .nav-bar { display: flex; justify-content: space-between; align-items: center; margin-bottom: 25px; flex-wrap: wrap; gap: 15px; }
            .nav-group { display: flex; gap: 10px; background-color: #121212; padding: 5px; border-radius: 10px; border: 1px solid #2A323C; }
            .nav-btn { background: none; border: none; color: #555555; font-size: 16px; font-weight: bold; cursor: pointer; padding: 8px 20px; border-radius: 6px; transition: all 0.3s; font-family: 'Consolas', monospace; }
            .nav-btn:hover { background-color: #2A323C; color: #FFFFFF !important; }
            .nav-btn.active { background-color: #2A323C; }
            
            .search-box { display: flex; gap: 10px; }
            .search-box input { background-color: #121212; border: 1px solid #2A323C; color: var(--theme-color); padding: 8px 15px; border-radius: 8px; font-family: 'Consolas', monospace; font-size: 16px; outline: none; transition: border-color 0.3s; width: 250px; }
            .search-box input:focus { border-color: var(--theme-color); }
            .search-box button { background-color: #1A1F24; border: 1px solid #2A323C; color: #FFFFFF; padding: 8px 15px; border-radius: 8px; cursor: pointer; transition: all 0.3s; font-family: 'Consolas', monospace; font-weight: bold; }
            .search-box button:hover { background-color: var(--theme-color); color: #121212; }

            .header { display: flex; justify-content: space-between; align-items: center; border-bottom: 3px solid var(--theme-color); padding-bottom: 20px; margin-bottom: 30px; transition: border-color 0.3s; }
            .header h1 { margin: 0; color: var(--theme-color); font-size: 32px; text-transform: uppercase; letter-spacing: 2px; transition: color 0.3s; min-width: 360px; }
            .header-info { display: flex; flex-direction: row; align-items: center; gap: 15px; }

            .top-stats { display: grid; grid-template-columns: repeat(4, 1fr); gap: 20px; margin-bottom: 40px; }
            .stat-box { background-color: #121212; border-radius: 12px; padding: 20px; text-align: center; border-left: 4px solid var(--theme-color); transition: border-color 0.3s; }
            .stat-box .title { font-size: 16px; color: #AAAAAA; margin-bottom: 8px; font-weight: bold; }
            .stat-box .value { font-size: 24px; font-weight: bold; color: #FFFFFF; font-family: 'Consolas', monospace; transition: text-shadow 0.3s, color 0.3s; }
            .stat-box .diff { font-size: 14px; color: var(--theme-color); transition: color 0.3s; }

            /* --- 🌟 排位賽專屬放大樣式 (僅放大文字，保持外框比例) --- */
            .stat-box.enlarged .title { font-size: 18px; margin-bottom: 6px; }
            .stat-box.enlarged .value { font-size: 32px; }
            .stat-box.enlarged .diff { font-size: 16px; }
            
            .summary-section { background-color: #121212; border-radius: 12px; padding: 25px; margin-bottom: 40px; }
            
            .brawler-grid { display: grid; grid-template-columns: 1fr 1fr; gap: 25px; }
            .brawler-cat { background-color: #121212; border-radius: 12px; padding: 20px; border: 1px solid #1A1F24; transition: all 0.3s ease; }
            .brawler-cat:hover { border-color: var(--theme-color); transform: translateY(-5px); }
            .brawler-cat h3 { margin: 0 0 15px 0; color: var(--theme-color); font-size: 20px; border-bottom: 2px solid #2A323C; padding-bottom: 12px; transition: color 0.3s; }
            
            .footer { text-align: center; margin-top: 30px; color: #555555; font-size: 14px; }

            .b-line { display: flex; justify-content: space-between; padding: 6px 0; font-family: 'Consolas', monospace; font-size: 16px; border-bottom: 1px solid #1A1F24; }
            .b-line:last-child { border-bottom: none; }
            .b-name { color: #DDDDDD; font-weight: bold; }
            .b-data { color: #FFFFFF; }
            
            .b-line-bar { padding: 6px 0; border-bottom: 1px solid #1A1F24; }
            .b-line-bar:last-child { border-bottom: none; }
            .b-line-bar .bar-label { display: flex; justify-content: space-between; margin-bottom: 5px; font-family: 'Consolas', monospace; font-size: 15px; }
            .b-line-bar .bar-track { display: flex; width: 100%; height: 6px; background-color: #2A323C; border-radius: 3px; overflow: hidden; }
            
            .summary-line { display: flex; justify-content: space-between; padding: 12px 0; border-bottom: 1px dashed #2A323C; font-size: 20px; font-family: 'Consolas', monospace; }
            .summary-line.is-total { border-bottom: none; font-weight: bold; font-size: 24px; color: var(--theme-color); margin-top: 10px; transition: color 0.3s; }
            
            .summary-line-bar { padding: 12px 0; border-bottom: 1px dashed #2A323C; }
            .summary-line-bar.is-total { border-bottom: none; margin-top: 10px; }
            .summary-line-bar .bar-label { display: flex; justify-content: space-between; margin-bottom: 8px; font-size: 20px; font-family: 'Consolas', monospace; }
            .summary-line-bar.is-total .bar-label { font-weight: bold; font-size: 24px; color: var(--theme-color); }
            .summary-line-bar .bar-track { display: flex; width: 100%; height: 8px; background-color: #2A323C; border-radius: 4px; overflow: hidden; }

            .bar-fill { height: 100%; transition: width 0.5s ease; }
            .bar-fill.win { background-color: var(--theme-color); }

            ::-webkit-scrollbar { width: 8px; }
            ::-webkit-scrollbar-track { background: transparent; }
            ::-webkit-scrollbar-thumb { background: rgba(255, 255, 255, 0.15); border-radius: 10px; border: 2px solid #1A1F24; }
            ::-webkit-scrollbar-thumb:hover { background: rgba(255, 255, 255, 0.3); }

            /* 模態彈窗樣式 */
            .modal { display: none; position: fixed; z-index: 1000; left: 0; top: 0; width: 100%; height: 100%; background-color: rgba(0,0,0,0.85); backdrop-filter: blur(5px); justify-content: center; align-items: center; }
            .modal-content { max-height: 95vh; overflow-y: auto; background-color: #1A1F24; width: 95%; max-width: 500px; border-radius: 15px; border: 1px solid #2A323C; box-shadow: 0 10px 40px rgba(0,0,0,0.8); transition: max-width 0.3s ease; }
            .modal-header { background: linear-gradient(135deg, #1A1F24 0%, #2A323C 100%); padding: 15px 25px; display: flex; justify-content: space-between; align-items: center; border-bottom: 3px solid var(--theme-color); position: sticky; top: 0; z-index: 10; }
            .modal-header h1 { margin: 0; color: #FFFFFF; font-size: 22px; font-family: 'Consolas', monospace; }
            .close-btn { color: #AAAAAA; font-size: 32px; font-weight: bold; cursor: pointer; transition: color 0.3s; line-height: 1; }
            .close-btn:hover { color: var(--theme-color); }
            .modal-body { padding: 15px 20px; }
            .modal-body .brawler-cat { margin-bottom: 12px; padding: 12px; }

            .map-view-grid { display: flex; flex-direction: column; gap: 12px; }
            .map-view-grid .brawler-cat { padding: 15px 25px; margin-bottom: 0; }
            .map-view-grid .brawler-cat h3 { margin: 0 0 5px 0; font-size: 20px; padding-bottom: 8px; }
            .map-view-grid .summary-line { padding: 8px 0; font-size: 18px; border-bottom: 1px solid #2A323C; margin-bottom: 6px; }
            .map-view-grid .b-line { padding: 5px 0; font-size: 16px; border-bottom: 1px dashed #1A1F24; }
            .map-view-grid .b-line:last-child { border-bottom: none; }
            
            .page-container { display: none; animation: fadeIn 0.3s ease-in-out; }
            .page-container.active { display: block; }
            @keyframes fadeIn { from { opacity: 0; transform: translateY(5px); } to { opacity: 1; transform: translateY(0); } }
        </style>
    </head>
    <body>
        <div class="container">
            <div class="nav-bar">
                <div class="nav-group" id="acc-nav">
                    <button class="nav-btn" onclick="switchAccount('大號')" id="btn-大號">大號</button>
                    <button class="nav-btn" onclick="switchAccount('小號')" id="btn-小號">小號</button>
                </div>
                
                <div class="nav-group" id="display-nav">
                    <button class="nav-btn" onclick="setDisplayMode('data')" id="btn-disp-data" title="文字數據版">🔢</button>
                    <button class="nav-btn" onclick="setDisplayMode('bar')" id="btn-disp-bar" title="進度條狀版">📊</button>
                </div>

                <div class="nav-group" id="align-nav">
                    <button class="nav-btn" onclick="setAlignment('flex-start')" id="btn-align-left" title="靠左對齊">⬅️</button>
                    <button class="nav-btn" onclick="setAlignment('center')" id="btn-align-center" title="置中對齊">⏹️</button>
                    <button class="nav-btn" onclick="setAlignment('flex-end')" id="btn-align-right" title="靠右對齊">➡️</button>
                </div>
                
                <div class="nav-group" id="view-nav">
                    <button class="nav-btn" onclick="switchView('session')" id="btn-session">▶ 本次區間</button>
                    <button class="nav-btn" onclick="switchView('all_time')" id="btn-all_time">▶ 歷史總計</button>
                </div>
            </div>

            <div class="header">
                <div class="header-info">
                    <h1 id="global-title">戰術主控台</h1>
                    <button id="btn-page-toggle" class="nav-btn" style="background-color: #2A323C; color: var(--theme-color); font-weight: bold; padding: 6px 15px; width: 150px; text-align: center;" onclick="togglePage()">▶ 排位賽專頁</button>
                </div>
                
                <div class="search-box" style="margin-bottom: 0;">
                    <input type="text" id="searchInput" placeholder="🔍 搜尋英雄、地圖或 ping" onkeypress="if(event.key === 'Enter') handleSearch()">
                    <button onclick="handleSearch()">查詢</button>
                </div>
            </div>

            <div id="page-main" class="page-container active">
                <div class="top-stats">
                    <div class="stat-box"><div class="title">🏆 總盃數</div><div class="value" id="val-trophies">- <span class="diff" id="diff-trophies">(-)</span></div></div>
                    <div class="stat-box"><div class="title">⚔️ 3V3 勝場</div><div class="value" id="val-3v3">-</div></div>
                    <div class="stat-box"><div class="title">🎯 排位 Elo</div><div class="value" id="val-elo">- <span class="diff" id="diff-elo">(-)</span></div></div>
                    <div class="stat-box"><div class="title">⭐ 排位段位</div><div class="value" id="val-tier">-</div></div>
                </div>
                
                <div class="summary-section" id="summary-section"></div>
                <div class="brawler-grid" id="brawler-grid"></div>
            </div>

            <div id="page-ranked" class="page-container">
                <div style="display: grid; grid-template-columns: repeat(3, 1fr); gap: 20px; margin-bottom: 30px;">
                    <div class="stat-box" style="text-align: left; display: flex; flex-direction: column; justify-content: center; gap: 8px;">
                        <div style="font-size:14px; color:#AAAAAA; font-weight:bold; margin-bottom:4px;">📶 即時網路監測</div>
                        <div style="display:flex; justify-content:space-between; font-family:'Consolas', monospace; font-size:16px;">
                            <span>官方 API:</span> <span id="val-ping-api" style="font-weight:bold;">--ms</span>
                        </div>
                        <div style="display:flex; justify-content:space-between; font-family:'Consolas', monospace; font-size:16px;">
                            <span>東京節點:</span> <span id="val-ping-tokyo" style="font-weight:bold;">--ms</span>
                        </div>
                        <div style="display:flex; justify-content:space-between; font-family:'Consolas', monospace; font-size:16px;">
                            <span>香港節點:</span> <span id="val-ping-hk" style="font-weight:bold;">--ms</span>
                        </div>
                    </div>
                    
                    <div class="stat-box enlarged" style="display: flex; flex-direction: column; justify-content: center;"><div class="title">🎯 排位 Elo</div><div class="value" id="val-elo-rk">- <span class="diff" id="diff-elo-rk">(-)</span></div></div>
                    <div class="stat-box enlarged" style="display: flex; flex-direction: column; justify-content: center;"><div class="title">⭐ 排位段位</div><div class="value" id="val-tier-rk">-</div></div>
                </div>
                
                <div class="summary-section" id="summary-ranked-only" style="margin-bottom: 40px; padding: 15px 25px;"></div>
                
                <div id="ranked-seasons-container"></div>
            </div>

            <div class="footer">伺服器自動生成於背景執行中... <br><span id="refresh-status" style="color:var(--theme-color);">畫面每 30 秒自動刷新最新戰況</span></div>
        </div>

        <div id="searchModal" class="modal">
            <div class="modal-content" id="modal-content-box">
                <div class="modal-header">
                    <h1 id="modal-title">戰術透視</h1>
                    <span class="close-btn" onclick="closeModal()">&times;</span>
                </div>
                <div class="modal-body" id="modal-body"></div>
            </div>
        </div>

        <script>
            let appData = __APP_DATA_HERE__;
            let currentPingData = __PING_DATA_HERE__;
            window.appData = appData;
            
            let currentAcc = sessionStorage.getItem('currentAcc') || "大號";
            let currentView = sessionStorage.getItem('currentView') || "session";
            let currentAlign = localStorage.getItem('pageAlign') || 'center';
            let currentDisplayMode = localStorage.getItem('displayMode') || 'data'; 
            let activePage = sessionStorage.getItem('activePage') || 'main';
            
            const TARGET_SIX_MODES = ['搶星大作戰', '寶石爭奪戰', '金庫攻防戰', '亂鬥足球', '據點搶奪戰', '極限淘汰賽'];

            function get_wr_js(w, l, d=0) {
                let total = w + l + d;
                return total > 0 ? (w/total*100).toFixed(1) + '%' : "0.0%";
            }

            function togglePage() {
                activePage = activePage === 'main' ? 'ranked' : 'main';
                sessionStorage.setItem('activePage', activePage);
                applyPageState();
                render();
            }

            function applyPageState() {
                document.getElementById('page-main').classList.toggle('active', activePage === 'main');
                document.getElementById('page-ranked').classList.toggle('active', activePage === 'ranked');
                
                const btn = document.getElementById('btn-page-toggle');
                const title = document.getElementById('global-title');
                
                if (btn && title) {
                    if (activePage === 'main') {
                        btn.innerHTML = '▶ 排位賽專頁';
                        title.innerText = `${currentAcc} 戰術主控台`;
                    } else {
                        btn.innerHTML = '◀ 返回主頁';
                        title.innerText = `${currentAcc} 排位賽深度解析`;
                    }
                }
            }

            function createRowHtml(label, statObj, isSummary = false, isTotal = false) {
                const total = statObj.w + statObj.l + statObj.d;
                let lineClass = isSummary ? 'summary-line' : 'b-line';
                if (isTotal) lineClass += ' is-total';
                
                const displayTxt = statObj.txt || statObj.stats;
                
                if (currentDisplayMode === 'data' || total === 0) {
                    return `<div class="${lineClass}"><span class="b-name">${label}</span><span class="b-data">${displayTxt}</span></div>`;
                } else {
                    const wPct = (statObj.w / total) * 100;
                    const barClass = isSummary ? 'summary-line-bar' + (isTotal ? ' is-total' : '') : 'b-line-bar';
                    
                    let trackHtml = `<div class="bar-track">`;
                    if(statObj.w > 0) trackHtml += `<div class="bar-fill win" style="width: ${wPct}%;"></div>`;
                    trackHtml += `</div>`;
                    
                    return `
                    <div class="${barClass}">
                        <div class="bar-label">
                            <span class="b-name">${label}</span>
                            <span class="b-data">${displayTxt}</span>
                        </div>
                        ${trackHtml}
                    </div>`;
                }
            }

            function renderRankedPage(accData) {
                const container = document.getElementById('ranked-seasons-container');
                if (!container) return;
                container.innerHTML = '';
                
                const isSession = (currentView === 'session');
                const seasonsData = isSession ? accData.ranked_seasons_session : accData.ranked_seasons_all_time;
                
                if (!seasonsData || Object.keys(seasonsData).length === 0) {
                    container.innerHTML = `
                        <div style="text-align:center; padding: 50px 20px; background-color:#121212; border-radius:12px; margin-top:20px; border:1px dashed #2A323C;">
                            <div style="font-size:32px; margin-bottom:10px;">${isSession ? '⏳' : '📊'}</div>
                            <div style="font-size:18px; color:#AAA; font-weight:bold;">${isSession ? '本次區間尚未進行任何排位賽' : '資料庫中尚無排位賽紀錄'}</div>
                            <div style="font-size:13px; color:#666; margin-top:5px;">${isSession ? '打完排位賽後將即時在此顯示本次實時對戰！' : ''}</div>
                        </div>`;
                    return;
                }

                const seasons = Object.keys(seasonsData).sort((a,b) => parseInt(b) - parseInt(a));
                
                seasons.forEach(season => {
                    const sData = seasonsData[season];
                    
                    let subBadge = isSession ? 
                        `<span style="font-size:14px; color:var(--theme-color); padding: 2px 8px; border: 1px solid var(--theme-color); border-radius: 4px; margin-left:10px;">本次對戰</span>` :
                        ((sData.start_date && sData.end_date) ? ` <span style="font-size:18px; color:#AAAAAA;">(${sData.start_date} ~ ${sData.end_date})</span>` : "");
                    
                    let sHtml = `<div class="season-section">
                        <h2 style="color:var(--theme-color); border-bottom: 2px solid #2A323C; padding-bottom: 10px; margin-top: 30px;">
                            🏆 第 ${season} 賽季${subBadge} <span style="font-size:16px; color:#888; float:right; line-height: 28px;">${sData.w}W - ${sData.l}L (${get_wr_js(sData.w, sData.l, sData.d)})</span>
                        </h2>
                        <div class="brawler-grid">`;
                    
                    const modeColors = {
                        '搶星大作戰': '#01cfff',
                        '寶石爭奪戰': '#9b3df3',
                        '金庫攻防戰': '#d65cd3',
                        '亂鬥足球': '#8ca0df',
                        '據點搶奪戰': '#e33c50',
                        '極限淘汰賽': '#f7831c'
                    };
                    
                    TARGET_SIX_MODES.forEach(modeName => {
                        let totalMatches = 0;
                        let brawlers = [];
                        
                        for (const [bName, bData] of Object.entries(sData.brawlers)) {
                            if (bData.modes && bData.modes[modeName]) {
                                let modeData = bData.modes[modeName];
                                let m = modeData.w + modeData.l + modeData.d;
                                totalMatches += m;
                                brawlers.push({ name: bName, w: modeData.w, l: modeData.l, d: modeData.d, matches: m, wr: modeData.w / m });
                            }
                        }
                        
                        brawlers.forEach(b => b.pr = totalMatches > 0 ? b.matches / totalMatches : 0);
                        
                        let color = modeColors[modeName] || '#FFFFFF';
                        let mHtml = `<div class="brawler-cat" style="border-top: 4px solid ${color};">
                            <h3 style="color: ${color}; margin-bottom: 5px;">${modeName}</h3>
                            <div style="text-align: right; font-size: 14px; color: #888; font-family: Consolas; margin-bottom: 15px;">總局數: ${totalMatches}</div>`;
                            
                        if (isSession) {
                            if (brawlers.length === 0) {
                                mHtml += `<div style="color:#777; text-align:center; padding: 30px 0;">(本次未出戰)</div>`;
                            } else {
                                mHtml += `<div style="color:#DDD; font-size:14px; margin: 10px 0 8px 0; font-weight:bold;">⚔️ 本次出戰英雄 (${totalMatches}場)</div>`;
                                brawlers.sort((a, b) => b.matches - a.matches || b.wr - a.wr);
                                brawlers.forEach(b => {
                                    mHtml += `<div class="b-line-bar"><div class="bar-label"><span class="b-name">🦸 ${b.name}</span><span class="b-data">${(b.wr*100).toFixed(1)}% (${b.w}W-${b.l}L)</span></div><div class="bar-track"><div class="bar-fill win" style="width: ${b.wr*100}%; background-color: ${color};"></div></div></div>`;
                                });
                            }
                        } else {
                            let valid = brawlers.filter(b => b.matches >= 3);
                            let topPR = [...valid].sort((a, b) => b.pr - a.pr).slice(0, 3);
                            let topWR = [...valid].sort((a, b) => b.wr - a.wr || b.matches - a.matches).slice(0, 3);
                            
                            let trap = [...valid].filter(b => b.wr < 0.45).sort((a, b) => b.matches - a.matches)[0];
                            let gem = [...valid].filter(b => b.wr >= 0.70 && !topPR.some(t => t.name === b.name)).sort((a, b) => b.wr - a.wr || b.matches - a.matches)[0];
                            
                            if (valid.length === 0) {
                                mHtml += `<div style="color:#777; text-align:center; padding: 30px 0;">(該模式需出場滿 3 次才能計算排行榜)</div>`;
                            } else {
                                mHtml += `<div style="color:#DDD; font-size:14px; margin: 10px 0 5px 0;">📊 出場率 Top 3</div>`;
                                topPR.forEach(b => {
                                    mHtml += `<div class="b-line-bar"><div class="bar-label"><span class="b-name">🦸 ${b.name}</span><span class="b-data">${(b.pr*100).toFixed(1)}% (${b.matches}場)</span></div><div class="bar-track"><div class="bar-fill" style="width: ${b.pr*100}%; background-color: #888888;"></div></div></div>`;
                                });
                                
                                mHtml += `<div style="color:#DDD; font-size:14px; margin: 20px 0 5px 0;">🏆 勝率 Top 3</div>`;
                                topWR.forEach(b => {
                                    mHtml += `<div class="b-line-bar"><div class="bar-label"><span class="b-name">🦸 ${b.name}</span><span class="b-data">${(b.wr*100).toFixed(1)}% (${b.w}W-${b.l}L)</span></div><div class="bar-track"><div class="bar-fill win" style="width: ${b.wr*100}%; background-color: ${color};"></div></div></div>`;
                                });
                                
                                if (trap || gem) {
                                    mHtml += `<div style="margin-top: 25px; padding: 12px; background-color: #1A1F24; border-radius: 6px; border-left: 3px solid #2A323C;">`;
                                    if (trap) mHtml += `<div style="margin-bottom: ${gem ? '12px' : '0'};"><div style="color:#FF5555; font-size:13px; font-weight:bold;">⚠️ 版本陷阱 (頭鐵掉分機)</div><div style="display:flex; justify-content:space-between; margin-top:5px; font-family:Consolas;"><span class="b-name">🦸 ${trap.name}</span><span style="color:#FFF;">${(trap.wr*100).toFixed(1)}% 勝率</span></div></div>`;
                                    if (gem) mHtml += `<div><div style="color:#00FFAA; font-size:13px; font-weight:bold;">💎 潛力神角 (上分奇兵)</div><div style="display:flex; justify-content:space-between; margin-top:5px; font-family:Consolas;"><span class="b-name">🦸 ${gem.name}</span><span style="color:#FFF;">${(gem.wr*100).toFixed(1)}% 勝率</span></div></div>`;
                                    mHtml += `</div>`;
                                }
                            }
                        }
                        mHtml += `</div>`;
                        sHtml += mHtml;
                    });
                    
                    sHtml += `</div></div>`; 
                    container.innerHTML += sHtml;
                });
            }

            function drawWebPingChart() {
                const canvas = document.getElementById('pingCanvas');
                if (!canvas) return;
                const ctx = canvas.getContext('2d');
                const w = canvas.width;
                const h = canvas.height;

                ctx.clearRect(0, 0, w, h);

                const hist = currentPingData.history;
                if (!hist || !hist.api || hist.api.length === 0) {
                    ctx.fillStyle = '#AAAAAA';
                    ctx.font = '14px Consolas';
                    ctx.textAlign = 'center';
                    ctx.fillText('📡 資料搜集中，請稍後...', w/2, h/2);
                    return;
                }

                let allVals = [...hist.api, ...hist.tokyo, ...hist.hk].filter(v => v < 999);
                let maxY = allVals.length > 0 ? Math.max(...allVals) : 100;
                maxY = Math.max(Math.ceil(maxY / 50) * 50, 100);

                ctx.strokeStyle = '#2A323C';
                ctx.fillStyle = '#777777';
                ctx.font = '12px Consolas';
                ctx.textAlign = 'left';

                for (let i = 1; i <= 5; i++) {
                    let y = h - (h * (i/5));
                    ctx.beginPath();
                    ctx.setLineDash([4, 4]);
                    ctx.moveTo(0, y);
                    ctx.lineTo(w, y);
                    ctx.stroke();
                    ctx.setLineDash([]);
                    ctx.fillText(Math.round(maxY * i / 5) + 'ms', 5, y - 5);
                }
                
                ctx.beginPath();
                ctx.strokeStyle = '#555555';
                ctx.lineWidth = 2;
                ctx.moveTo(0, h);
                ctx.lineTo(w, h);
                ctx.stroke();

                function drawLine(data, color) {
                    if (!data || data.length === 0) return;
                    ctx.strokeStyle = color;
                    ctx.lineWidth = 2;
                    ctx.beginPath();

                    let xStep = w / 60;
                    let startX = w - (data.length * xStep);

                    data.forEach((val, i) => {
                        let x = startX + (i * xStep);
                        let drawVal = val < 999 ? val : maxY * 1.1;
                        let y = h - (h * (drawVal / maxY));
                        y = Math.max(0, Math.min(y, h));

                        if (i === 0) ctx.moveTo(x, y);
                        else ctx.lineTo(x, y);
                    });
                    ctx.stroke(); 

                    let lastVal = data[data.length - 1];
                    let drawVal = lastVal < 999 ? lastVal : maxY * 1.1;
                    let lastY = h - (h * (drawVal / maxY));
                    lastY = Math.max(0, Math.min(lastY, h));
                    let lastX = startX + ((data.length - 1) * xStep);

                    ctx.fillStyle = color;
                    ctx.beginPath();
                    ctx.arc(lastX, lastY, 3, 0, 2 * Math.PI);
                    ctx.fill();
                }

                drawLine(hist.api, '#00CCFF');
                drawLine(hist.tokyo, '#DF44FF');
                drawLine(hist.hk, '#FFD700');
            }

            function updatePingUI(pingData) {
                const pingApiEl = document.getElementById('val-ping-api');
                const pingTokyoEl = document.getElementById('val-ping-tokyo');
                const pingHkEl = document.getElementById('val-ping-hk');
                
                if (pingApiEl) {
                    let apiColor = pingData.api < 80 ? '#00FFAA' : (pingData.api < 150 ? '#FFAA00' : '#FF5555');
                    pingApiEl.innerText = pingData.api === 999 ? '異常' : `${pingData.api}ms`;
                    pingApiEl.style.color = apiColor;
                }
                if (pingTokyoEl) {
                    let tokyoColor = pingData.tokyo < 80 ? '#00FFAA' : (pingData.tokyo < 150 ? '#FFAA00' : '#FF5555');
                    pingTokyoEl.innerText = pingData.tokyo === 999 ? '異常' : `${pingData.tokyo}ms`;
                    pingTokyoEl.style.color = tokyoColor;
                }
                if (pingHkEl) {
                    let hkColor = pingData.hk < 80 ? '#00FFAA' : (pingData.hk < 150 ? '#FFAA00' : '#FF5555');
                    pingHkEl.innerText = pingData.hk === 999 ? '異常' : `${pingData.hk}ms`;
                    pingHkEl.style.color = hkColor;
                }
            }

            function render() {
                const data = appData[currentAcc];
                const viewData = data[currentView];
                const isSession = (currentView === 'session');
                
                document.documentElement.style.setProperty('--theme-color', data.color);
                applyPageState();
                
                updatePingUI(currentPingData.current);
                
                ['btn-大號', 'btn-小號', 'btn-session', 'btn-all_time', 'btn-disp-data', 'btn-disp-bar'].forEach(id => {
                    const el = document.getElementById(id);
                    if(!el) return;
                    if(id.includes('大號')) { el.classList.toggle('active', currentAcc === '大號'); el.style.color = currentAcc === '大號' ? data.color : '#555555'; }
                    if(id.includes('小號')) { el.classList.toggle('active', currentAcc === '小號'); el.style.color = currentAcc === '小號' ? data.color : '#555555'; }
                    if(id.includes('session')) { el.classList.toggle('active', isSession); el.style.color = isSession ? data.color : '#555555'; }
                    if(id.includes('all_time')) { el.classList.toggle('active', !isSession); el.style.color = !isSession ? data.color : '#555555'; }
                    if(id.includes('data')) { el.classList.toggle('active', currentDisplayMode === 'data'); el.style.color = currentDisplayMode === 'data' ? data.color : '#555555'; }
                    if(id.includes('bar')) { el.classList.toggle('active', currentDisplayMode === 'bar'); el.style.color = currentDisplayMode === 'bar' ? data.color : '#555555'; }
                });
                
                const tierStr = data.tier.toUpperCase();
                let tierColor = data.color; 
                if (tierStr.includes('BRONZE')) tierColor = '#CD7F32';      
                else if (tierStr.includes('SILVER')) tierColor = '#B4C5E4'; 
                else if (tierStr.includes('GOLD')) tierColor = '#FFD700';   
                else if (tierStr.includes('DIAMOND')) tierColor = '#11C4EB';
                else if (tierStr.includes('MYTHIC')) tierColor = '#DF44FF'; 
                else if (tierStr.includes('LEGENDARY')) tierColor = '#FF3333'; 
                else if (tierStr.includes('MASTER')) tierColor = '#FF8800'; 
                else if (tierStr.includes('PRO')) tierColor = '#33CC33';    

                const valTrophies = document.getElementById('val-trophies');
                if(valTrophies) valTrophies.innerHTML = `${data.trophies} <span class="diff">(${data.diff_trophies})</span>`;
                
                const val3v3 = document.getElementById('val-3v3');
                if(val3v3) val3v3.innerText = data.victories_3v3;
                
                const valElo = document.getElementById('val-elo');
                if(valElo) valElo.innerHTML = `${data.elo} <span class="diff">(${data.diff_elo})</span>`;
                
                const tierElem = document.getElementById('val-tier');
                if(tierElem) {
                    tierElem.innerText = data.tier;
                    tierElem.style.color = tierColor;
                    tierElem.style.textShadow = `0 0 15px ${tierColor}90`;
                }
                
                const sumSec = document.getElementById('summary-section');
                if(sumSec) {
                    sumSec.innerHTML = `
                        ${createRowHtml('🏅 排位賽', viewData.summary.ranked, true)}
                        ${createRowHtml('⏳ 一般模式', viewData.summary.casual, true)}
                        ${createRowHtml('🎪 特別活動', viewData.summary.special, true)}
                        ${createRowHtml('📊 總戰績', viewData.summary.total, true, true)}
                    `;
                }
                
                const grid = document.getElementById('brawler-grid');
                if(grid) {
                    grid.innerHTML = '';
                    viewData.brawlers.forEach(cat => {
                        let catHtml = `<div class="brawler-cat"><h3>${cat.icon} ${cat.title}</h3>`;
                        cat.items.forEach(b => {
                            catHtml += createRowHtml(`🦸 ${b.name}`, b);
                        });
                        catHtml += `</div>`;
                        grid.innerHTML += catHtml;
                    });
                }

                const valEloRk = document.getElementById('val-elo-rk');
                if(valEloRk) valEloRk.innerHTML = `${data.elo} <span class="diff">(${data.diff_elo})</span>`;
                
                const valTierRk = document.getElementById('val-tier-rk');
                if(valTierRk) {
                    valTierRk.innerText = data.tier;
                    valTierRk.style.color = tierColor;
                    valTierRk.style.textShadow = `0 0 15px ${tierColor}90`;
                }
                
                const sumSecRk = document.getElementById('summary-ranked-only');
                if(sumSecRk) {
                    const rkLabel = isSession ? '🏅 排位戰績 (本次)' : '🏅 排位總計 (歷史)';
                    sumSecRk.innerHTML = createRowHtml(rkLabel, viewData.summary.ranked, true);
                }
                
                renderRankedPage(data);
                
                const searchModal = document.getElementById('searchModal');
                if (searchModal && searchModal.style.display === 'flex') {
                    if (!document.getElementById('pingCanvas')) {
                        handleSearch(true);
                    }
                }
            }

            function switchAccount(acc) {
                currentAcc = acc;
                sessionStorage.setItem('currentAcc', acc);
                render();
            }
            
            function switchView(view) {
                currentView = view;
                sessionStorage.setItem('currentView', view);
                render();
            }

            function setDisplayMode(mode) {
                currentDisplayMode = mode;
                localStorage.setItem('displayMode', mode);
                render();
            }
            
            function setAlignment(align) {
                currentAlign = align;
                localStorage.setItem('pageAlign', align);
                document.body.style.justifyContent = align;
                
                const btnAlignLeft = document.getElementById('btn-align-left');
                if(btnAlignLeft) btnAlignLeft.classList.toggle('active', align === 'flex-start');
                
                const btnAlignCenter = document.getElementById('btn-align-center');
                if(btnAlignCenter) btnAlignCenter.classList.toggle('active', align === 'center');
                
                const btnAlignRight = document.getElementById('btn-align-right');
                if(btnAlignRight) btnAlignRight.classList.toggle('active', align === 'flex-end');
            }

            function handleSearch(isReRender = false) {
                const searchInput = document.getElementById('searchInput');
                if(!searchInput) return;
                const query = searchInput.value.trim();
                if (!query && !isReRender) return;
                
                if (query.toLowerCase() === 'ping') {
                    document.getElementById('modal-title').innerText = "戰術透視鏡 - 網路雷達";
                    const modalBox = document.getElementById('modal-content-box');
                    if (modalBox) modalBox.style.maxWidth = "540px";
                    
                    let html = `
                        <div style="text-align:center; padding: 10px;">
                            <div style="font-size:16px; font-weight:bold; margin-bottom:15px; color:#FFFFFF;">📶 過去 3 分鐘 Ping 值走勢</div>
                            <canvas id="pingCanvas" width="460" height="250" style="background-color:#1A1F24; border:1px solid #2A323C; border-radius:8px;"></canvas>
                            <div style="margin-top:15px; display:flex; justify-content:center; gap:20px; font-family:'Consolas', monospace; font-size:14px; font-weight:bold;">
                                <span style="color:#00CCFF;">■ 官方 API</span>
                                <span style="color:#DF44FF;">■ 東京節點</span>
                                <span style="color:#FFD700;">■ 香港節點</span>
                            </div>
                        </div>
                    `;
                    document.getElementById('modal-body').innerHTML = html;
                    document.getElementById('searchModal').style.display = 'flex';
                    document.body.classList.add('no-scroll');
                    
                    drawWebPingChart(); 
                    
                    if (!isReRender) searchInput.value = '';
                    return;
                }

                const isChinese = /[\\u4e00-\\u9fff]/.test(query);
                const searchData = appData[currentAcc]['all_time'];
                let resultHtml = "";
                let modalTitle = "";
                const modalBox = document.getElementById('modal-content-box');

                if (isChinese) {
                    modalTitle = "【 全模式地圖勝率 (歷史總計) 】";
                    if (modalBox) modalBox.style.maxWidth = "500px"; 
                    
                    resultHtml += `<div class="map-view-grid">`;
                    const mapCategories = [['🏅', '排位賽'], ['⏳', '一般模式']]; 
                    mapCategories.forEach(([icon, cat]) => {
                        let catData = searchData.map_stats.find(c => c.title === cat);
                        if (catData) {
                            resultHtml += `<div class="brawler-cat"><h3>${icon} ${cat} <span style="float:right; color:var(--theme-color); font-family:Consolas;">${catData.wr}</span></h3>`;
                            resultHtml += createRowHtml('分類總計', {stats: `${catData.wins}W - ${catData.losses}L`, w: catData.w, l: catData.l, d: catData.d}, true);
                            catData.modes.forEach(m => {
                                resultHtml += createRowHtml(`• ${m.name}`, m);
                            });
                            resultHtml += `</div>`;
                        }
                    });
                    resultHtml += `</div>`;
                } else {
                    if (modalBox) modalBox.style.maxWidth = "500px";
                    const bName = Object.keys(searchData.brawler_details).find(k => k.includes(query.toUpperCase()));
                    if (!bName) {
                        if(!isReRender) alert(`資料庫中找不到包含【${query}】的英雄紀錄。`);
                        return;
                    }
                    const bStats = searchData.brawler_details[bName];
                    modalTitle = `【 ${bName} 】(歷史總計)`;
                    
                    let totalRankedMatches = searchData.summary.ranked.w + searchData.summary.ranked.l + searchData.summary.ranked.d;

                    resultHtml += `<div class="brawler-cat" style="border-left-color:#FFAA00;"><h3>▶ 總結 <span style="float:right; color:#FFAA00; font-family:Consolas;">${bStats.summary.split('(')[1].replace(')','')}</span></h3>`;
                    resultHtml += createRowHtml('總勝負', {stats: bStats.summary.split('(')[0].trim(), w: bStats.w, l: bStats.l, d: bStats.d});
                    resultHtml += `</div>`;
                    
                    bStats.cats.forEach(cat => {
                        let prText = "";
                        if (cat.title === '排位賽') {
                            let catMatches = cat.w + cat.l + cat.d;
                            let pr = totalRankedMatches > 0 ? ((catMatches / totalRankedMatches) * 100).toFixed(1) : "0.0";
                            prText = ` <span style="font-size:14px; color:#888;">(出場率: ${pr}%)</span>`;
                        }
                        
                        resultHtml += `<div class="brawler-cat"><h3>${cat.icon} ${cat.title}${prText} <span style="float:right; color:var(--theme-color); font-family:Consolas;">${cat.wr}</span></h3>`;
                        resultHtml += createRowHtml('分類總計', {stats: `${cat.wins}W - ${cat.losses}L`, w: cat.w, l: cat.l, d: cat.d});
                        cat.modes.forEach(m => {
                            resultHtml += createRowHtml(`• ${m.name}`, m);
                        });
                        resultHtml += `</div>`;
                    });
                }
                
                const mTitle = document.getElementById('modal-title');
                if(mTitle) mTitle.innerText = modalTitle;
                
                const mBody = document.getElementById('modal-body');
                if(mBody) mBody.innerHTML = resultHtml;
                
                const mModal = document.getElementById('searchModal');
                if(mModal) mModal.style.display = 'flex';
                
                document.body.classList.add('no-scroll');
            }

            function closeModal() {
                const mModal = document.getElementById('searchModal');
                if(mModal) mModal.style.display = 'none';
                
                const rStatus = document.getElementById('refresh-status');
                if(rStatus) rStatus.innerText = "畫面每 30 秒自動刷新最新戰況";
                
                document.body.classList.remove('no-scroll');
            }

            setInterval(() => {
                const mModal = document.getElementById('searchModal');
                if (mModal && mModal.style.display !== 'flex') {
                    const script = document.createElement('script');
                    script.src = 'dashboard_data.js?t=' + new Date().getTime();
                    script.onload = function() {
                        try {
                            if (window.__DYNAMIC_APP_DATA__) {
                                appData = window.__DYNAMIC_APP_DATA__;
                                window.appData = appData;
                                render(); 
                            }
                        } catch (e) {}
                        document.body.removeChild(script); 
                    };
                    script.onerror = function() { document.body.removeChild(script); };
                    document.body.appendChild(script);
                } else {
                    const rStatus = document.getElementById('refresh-status');
                    if(rStatus && !document.getElementById('pingCanvas')) rStatus.innerText = "(為避免干擾閱讀，資料更新已暫停，關閉彈窗後恢復更新)";
                }
            }, 30000);
            
            setInterval(() => {
                const isPingModalOpen = document.getElementById('searchModal').style.display === 'flex' && document.getElementById('pingCanvas');
                
                if (activePage !== 'ranked' && !isPingModalOpen) return;
                
                const script = document.createElement('script');
                script.src = 'ping_data.js?t=' + new Date().getTime();
                script.onload = function() {
                    try {
                        if (window.__PING_DATA__) {
                            currentPingData = window.__PING_DATA__;
                            updatePingUI(currentPingData.current);
                            if (isPingModalOpen) drawWebPingChart(); 
                        }
                    } catch (e) {}
                    document.body.removeChild(script); 
                };
                script.onerror = function() { document.body.removeChild(script); };
                document.body.appendChild(script);
            }, 3000);

            render();
            setAlignment(currentAlign);
        </script>
    </body>
    </html>
    """
    
    html_path = os.path.join(base_dir, "Brawl_Tactics_Dashboard.html")
    try:
        final_html = html_template.replace('__APP_DATA_HERE__', js_string)
        final_html = final_html.replace('__PING_DATA_HERE__', js_ping_string)
        with open(html_path, 'w', encoding='utf-8') as f:
            f.write(final_html)
            
        data_js_path = os.path.join(base_dir, "dashboard_data.js")
        with open(data_js_path, 'w', encoding='utf-8') as f:
            f.write(f"window.__DYNAMIC_APP_DATA__ = {js_string};")
    except: pass

def brawlboard_keepalive_worker():
    while True:
        for acc_name, acc_info in ACCOUNTS.items():
            try:
                clean_tag = acc_info['tag'].replace('#', '')
                url = f"https://brawlboard.net/dc-account/{clean_tag}"
                requests.get(url, headers=headers_ninja, timeout=10)
            except:
                pass
        time.sleep(180)

def ping_worker():
    global network_ping_api, network_ping_tokyo, network_ping_hk
    while True:
        try:
            start = time.time()
            sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            sock.settimeout(2)
            sock.connect(('api.brawlstars.com', 443))
            sock.close()
            network_ping_api = int((time.time() - start) * 1000)
        except: network_ping_api = 999
            
        try:
            start = time.time()
            sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            sock.settimeout(2)
            sock.connect(('ec2.ap-northeast-1.amazonaws.com', 443))
            sock.close()
            network_ping_tokyo = int((time.time() - start) * 1000)
        except: network_ping_tokyo = 999
            
        try:
            start = time.time()
            sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            sock.settimeout(2)
            sock.connect(('ec2.ap-east-1.amazonaws.com', 443))
            sock.close()
            network_ping_hk = int((time.time() - start) * 1000)
        except: network_ping_hk = 999
            
        ping_history['api'].append(network_ping_api)
        ping_history['tokyo'].append(network_ping_tokyo)
        ping_history['hk'].append(network_ping_hk)
            
        ping_js_path = os.path.join(base_dir, "ping_data.js")
        try:
            ping_payload = {
                'current': {'api': network_ping_api, 'tokyo': network_ping_tokyo, 'hk': network_ping_hk},
                'history': {
                    'api': list(ping_history['api']),
                    'tokyo': list(ping_history['tokyo']),
                    'hk': list(ping_history['hk'])
                }
            }
            with open(ping_js_path, 'w', encoding='utf-8') as f:
                f.write(f"window.__PING_DATA__ = {json.dumps(ping_payload)};")
        except: pass
            
        time.sleep(3)

def fetch_account_data(acc_name):
    global account_stats, current_display_account, current_view_mode, last_sync_time_str, last_new_data_real_time
    
    # 🚀 錯開請求：改為讓大號晚 2 秒出發，小號先拿資料
    if acc_name == '大號':
        time.sleep(2)
        
    tag = ACCOUNTS[acc_name]['tag']
    excel_file = ACCOUNTS[acc_name]['excel']
    encoded_tag = urllib.parse.quote(tag)
    p_url = f'https://api.brawlstars.com/v1/players/{encoded_tag}'
    b_url = f'https://api.brawlstars.com/v1/players/{encoded_tag}/battlelog'
    
    while True:
        try:
            stats = account_stats[acc_name]
            res_profile = requests.get(p_url, headers=headers_official, timeout=10)
            res_battlelog = requests.get(b_url, headers=headers_official, timeout=10)
            
            if res_profile.status_code == 200 and res_battlelog.status_code == 200:
                last_sync_time_str = (datetime.utcnow() + timedelta(hours=8)).strftime('%H:%M:%S')
                data = res_profile.json()
                stats['owned_brawlers'] = [b.get('name', '').upper() for b in data.get('brawlers', [])]
                current_trophies = data.get('trophies', 0)
                if stats['start_trophies'] is None: stats['start_trophies'] = current_trophies
                trophy_diff = current_trophies - stats['start_trophies']
                stats['current_trophies'] = current_trophies
                stats['diff_str'] = f"+{trophy_diff}" if trophy_diff >= 0 else str(trophy_diff)
                stats['3v3_victories'] = data.get('3vs3Victories', 0)
                current_elo_val = data.get('rankedElo', 0)
                
                current_season_id = data.get('rankedSeasonId', 48)
                
                if current_elo_val > 0:
                    if stats['start_elo'] is None: stats['start_elo'] = current_elo_val
                    elo_diff = current_elo_val - stats['start_elo']
                    stats['elo_str'] = str(current_elo_val)
                    stats['elo_diff_str'] = f"+{elo_diff}" if elo_diff >= 0 else str(elo_diff)
                    stats['elo_tier'] = data.get('rankedRankName', 'UNKNOWN')

                excel_needs_update = False
                data_battle = res_battlelog.json()
                battles = data_battle.get('items', [])
                if battles:
                    latest_battle = battles[0]
                    current_battle_time = latest_battle.get('battleTime')
                    if stats['last_raw_time'] != current_battle_time:
                        stats['last_raw_time'] = current_battle_time
                        time_daho = account_stats['大號']['last_raw_time']
                        time_xiaho = account_stats['小號']['last_raw_time']
                        if time_daho and time_xiaho: current_display_account = '大號' if time_daho > time_xiaho else '小號'
                        elif time_daho: current_display_account = '大號'
                        elif time_xiaho: current_display_account = '小號'
                    if stats['last_time'] is None or current_battle_time != stats['last_time']:
                        excel_needs_update = True 
                        stats['last_time'] = current_battle_time

                    if excel_needs_update:
                        last_new_data_real_time = time.time()
                        
                        new_records = []
                        current_owned = stats.get('owned_brawlers', [])
                        for item in battles:
                            battle = item.get('battle', {})
                            event = item.get('event', {})
                            raw_time = item.get('battleTime', '')
                            formatted_time = raw_time
                            if len(raw_time) >= 15:
                                try:
                                    dt = datetime.strptime(raw_time[:15], '%Y%m%dT%H%M%S')
                                    dt_local = dt + timedelta(hours=8)
                                    formatted_time = dt_local.strftime('%Y-%m-%d %H:%M:%S') 
                                except: pass
                            raw_mode = event.get('mode') or battle.get('mode', '')
                            map_name = event.get('map', '')
                            raw_type = battle.get('type', '') 
                            raw_result = battle.get('result', '')
                            t_change = battle.get('trophyChange')
                            tc_excel = '無機制' if t_change is None else t_change
                            my_brawler = '未知'
                            my_brawler_trophies = None
                            my_team_brawlers = []
                            enemy_brawlers = []
                            teams = battle.get('teams', [])
                            players = battle.get('players', [])
                            if teams:
                                for team in teams:
                                    is_my_team = any(p.get('tag') == tag for p in team)
                                    brawler_names = [p.get('brawler', {}).get('name', '') for p in team]
                                    if is_my_team:
                                        my_team_brawlers = brawler_names
                                        for p in team:
                                            if p.get('tag') == tag:
                                                my_brawler = p.get('brawler', {}).get('name', '未知')
                                                my_brawler_trophies = p.get('brawler', {}).get('trophies')
                                    else: enemy_brawlers.extend(brawler_names)
                            elif players:
                                for p in players:
                                    b_name = p.get('brawler', {}).get('name', '')
                                    if p.get('tag') == tag:
                                        my_brawler = b_name
                                        my_brawler_trophies = p.get('brawler', {}).get('trophies')
                                        my_team_brawlers = [b_name]
                                    else: enemy_brawlers.append(b_name)
                            star_player = battle.get('starPlayer') or {}
                            is_mvp = bool(star_player.get('tag') == tag)
                            
                            rank_mech = '一般'
                            rank_season = ''
                            if raw_type in ['soloRanked', 'teamRanked']:
                                rank_season = str(current_season_id)
                                try:
                                    dt_val = datetime.strptime(formatted_time, '%Y-%m-%d %H:%M:%S')
                                    is_old_record = dt_val <= datetime(2026, 8, 19, 23, 59, 59)
                                except:
                                    is_old_record = False
                                    
                                if is_old_record:
                                    rank_mech = 'BO3'
                                    rank_season = '47'
                                elif my_brawler_trophies is None or str(my_brawler_trophies).strip() == '':
                                    rank_mech = 'BO3'
                                else:
                                    try:
                                        if int(my_brawler_trophies) >= 13: rank_mech = 'BO3'
                                        else: rank_mech = 'BO1'
                                    except: rank_mech = 'BO3'

                            new_records.append({
                                '對戰時間': formatted_time, '模式': raw_mode, '地圖': map_name, '類型': raw_type,  
                                '我方英雄': my_brawler, '英雄盃數': my_brawler_trophies, '戰果': raw_result,
                                '獎盃變化': tc_excel, '是否MVP': is_mvp, '我方陣容': ', '.join(my_team_brawlers),
                                '敵方陣容': ', '.join(enemy_brawlers), '排位機制': rank_mech, '賽季': rank_season
                            })

                        df_new = pd.DataFrame(new_records)
                        
                        read_success = True
                        df_old = pd.DataFrame()
                        if os.path.exists(excel_file):
                            try:
                                xls = pd.ExcelFile(excel_file)
                                sheet_to_read = '戰績明細' if '戰績明細' in xls.sheet_names else xls.sheet_names[0]
                                df_old = pd.read_excel(xls, sheet_name=sheet_to_read)
                                if not df_old.empty and '對戰時間' in df_old.columns:
                                    df_old['對戰時間'] = df_old['對戰時間'].astype(str)
                            except Exception as e:
                                read_success = False
                                print(f"讀取 Excel 失敗，啟動防護機制暫停寫入: {e}")
                                
                        if not read_success:
                            time.sleep(30)
                            continue

                        if not df_old.empty:
                            df_old['我方英雄'] = df_old['我方英雄'].fillna('未知')
                            if '是否寶箱活動' not in df_old.columns: df_old['是否寶箱活動'] = False
                            else: df_old['是否寶箱活動'] = df_old['是否寶箱活動'].fillna(False).astype(bool)
                            
                            if '賽季' not in df_old.columns: df_old['賽季'] = ''
                            
                            def correct_rank_mech(r):
                                if r.get('類型') in ['soloRanked', 'teamRanked']:
                                    dt_str = str(r.get('對戰時間', ''))
                                    try:
                                        if datetime.strptime(dt_str, '%Y-%m-%d %H:%M:%S') <= datetime(2026, 8, 19, 23, 59, 59):
                                            return 'BO3'
                                    except: pass
                                    
                                    val = r.get('英雄盃數')
                                    if pd.isna(val) or str(val).strip() == '':
                                        return 'BO3'
                                    try: 
                                        return 'BO3' if int(val) >= 13 else 'BO1'
                                    except: 
                                        return 'BO3'
                                return '一般'
                            
                            def determine_season(r):
                                if r.get('類型') not in ['soloRanked', 'teamRanked']: return ''
                                val = r.get('賽季')
                                if pd.notna(val) and str(val).strip() != '': return str(val)
                                dt_str = str(r.get('對戰時間', ''))
                                try:
                                    if datetime.strptime(dt_str, '%Y-%m-%d %H:%M:%S') <= datetime(2026, 8, 19, 23, 59, 59): 
                                        return '47'
                                    return str(current_season_id)
                                except: return '48'

                            df_old['排位機制'] = df_old.apply(correct_rank_mech, axis=1)
                            df_old['賽季'] = df_old.apply(determine_season, axis=1)

                        if not df_new.empty:
                            df_new['對戰時間'] = df_new['對戰時間'].astype(str)
                            df_new['是否寶箱活動'] = False
                            if current_owned:
                                def mark_new_boxes(r):
                                    try:
                                        t1 = str(r.get('我方陣容', ''))
                                        t2 = str(r.get('敵方陣容', ''))
                                        b_list = [x.strip() for x in (t1 + ',' + t2).split(',') if x.strip() and x.strip() != 'nan']
                                        if len(b_list) > 1 and len(set(b_list)) == 1: return False
                                        rtype = str(r.get('類型', ''))
                                        hero = str(r.get('我方英雄', '')).upper()
                                        if rtype == 'ranked' and hero not in current_owned and hero not in ['未知', 'NAN', 'NONE', '']: return True
                                        return False
                                    except: return False
                                df_new['是否寶箱活動'] = df_new.apply(mark_new_boxes, axis=1)

                        if not df_old.empty and not df_new.empty: 
                            df_combined = pd.concat([df_new, df_old], ignore_index=True).drop_duplicates(subset=['對戰時間'], keep='first')
                        elif not df_new.empty: df_combined = df_new
                        else: df_combined = df_old

                        if '是否寶箱活動' in df_combined.columns: df_combined['是否寶箱活動'] = df_combined['是否寶箱活動'].fillna(False).astype(bool)
                        df_combined = df_combined.sort_values(by='對戰時間', ascending=False)
                        
                        df_all_time_grouped = process_and_group_dataframe(df_combined)
                        stats['ui_all_time'] = build_ui_dict(df_all_time_grouped)
                        stats['ranked_seasons_all_time'] = build_ranked_ui_dict(df_all_time_grouped)
                        
                        df_session = df_combined[df_combined['對戰時間'] > stats['startup_formatted_time']].copy()
                        df_session_grouped = process_and_group_dataframe(df_session)
                        stats['ui_session'] = build_ui_dict(df_session_grouped)
                        stats['ranked_seasons_session'] = build_ranked_ui_dict(df_session_grouped)

                        summary_list = []
                        if not df_all_time_grouped.empty:
                            for brawler, group in df_all_time_grouped.groupby('我方英雄'):
                                total_matches = len(group)
                                wins = len(group[group['戰果'] == 'victory'])
                                losses = len(group[group['戰果'] == 'defeat'])
                                draws = len(group[group['戰果'] == 'draw'])
                                win_rate = wins / total_matches if total_matches > 0 else 0
                                summary_list.append({'英雄名稱': brawler, '總場數': total_matches, '勝場': wins, '敗場': losses, '平局': draws, '勝率': f"{win_rate:.2%}"})
                        df_summary = pd.DataFrame(summary_list)
                        if not df_summary.empty:
                            df_summary = df_summary.sort_values(by=['英雄名稱'], ascending=True)
                            total_matches_all = len(df_all_time_grouped)
                            total_wins_all = len(df_all_time_grouped[df_all_time_grouped['戰果'] == 'victory'])
                            total_losses_all = len(df_all_time_grouped[df_all_time_grouped['戰果'] == 'defeat'])
                            total_draws_all = len(df_all_time_grouped[df_all_time_grouped['戰果'] == 'draw'])
                            total_win_rate = total_wins_all / total_matches_all if total_matches_all > 0 else 0
                            total_row = pd.DataFrame([{'英雄名稱': '總計 (TOTAL)', '總場數': total_matches_all, '勝場': total_wins_all, '敗場': total_losses_all, '平局': total_draws_all, '勝率': f"{total_win_rate:.2%}"}])
                            df_summary = pd.concat([df_summary, total_row], ignore_index=True)

                        mode_summary_list = []
                        if not df_all_time_grouped.empty:
                            for (brawler, raw_t, raw_m), group in df_all_time_grouped.groupby(['我方英雄', 'UI動態分類', '模式中文']):
                                m_total = len(group)
                                m_wins = len(group[group['戰果'] == 'victory'])
                                m_losses = len(group[group['戰果'] == 'defeat'])
                                m_draws = len(group[group['戰果'] == 'draw'])
                                m_win_rate = m_wins / m_total if m_total > 0 else 0
                                mode_summary_list.append({'英雄名稱': brawler, '賽事類型': raw_t, '對戰模式': raw_m, '該模式場數': m_total, '勝場': m_wins, '敗場': m_losses, '平局': m_draws, '該模式勝率': f"{m_win_rate:.2%}"})
                        df_mode_summary = pd.DataFrame(mode_summary_list)
                        if not df_mode_summary.empty:
                            type_order = {'排位賽': 1, '一般模式': 2, '挑戰': 3, '寶箱活動': 4, '特別活動': 5, '鏡像亂鬥': 6}
                            df_mode_summary['排序權重'] = df_mode_summary['賽事類型'].map(type_order).fillna(99)
                            df_mode_summary = df_mode_summary.sort_values(by=['英雄名稱', '排序權重', '該模式場數'], ascending=[True, True, False])
                            df_mode_summary = df_mode_summary.drop(columns=['排序權重'])

                        # 🚀 即刻更新 Web UI 資料檔案
                        generate_interactive_main_page(current_display_account, current_view_mode)

                        if not df_combined.empty and not df_summary.empty:
                            try:
                                backup_file = excel_file.replace('.xlsx', f'_backup_{datetime.now().strftime("%Y%m%d")}.xlsx')
                                if os.path.exists(excel_file) and not os.path.exists(backup_file):
                                    shutil.copy2(excel_file, backup_file)
                                
                                temp_excel_file = excel_file.replace('.xlsx', '_tmp.xlsx')
                                
                                with pd.ExcelWriter(temp_excel_file, engine='openpyxl') as writer:
                                    df_combined.to_excel(writer, index=False, sheet_name='戰績明細')
                                    df_summary.to_excel(writer, index=False, sheet_name='英雄總勝率')
                                    df_mode_summary.to_excel(writer, index=False, sheet_name='模式詳細勝率') 
                                    
                                    center_alignment = Alignment(horizontal='center', vertical='center')
                                    for sheet_name in writer.sheets:
                                        worksheet = writer.sheets[sheet_name]
                                        worksheet.auto_filter.ref = worksheet.dimensions
                                        for i in range(1, 28): worksheet.column_dimensions[get_column_letter(i)].width = 22
                                        if sheet_name in ['英雄總勝率', '模式詳細勝率']:
                                            for row in worksheet.iter_rows():
                                                for cell in row: cell.alignment = center_alignment
                                
                                os.replace(temp_excel_file, excel_file)
                                
                            except PermissionError:
                                if os.path.exists(temp_excel_file):
                                    try: os.remove(temp_excel_file)
                                    except: pass
                            except Exception as e:
                                print(f"存檔發生嚴重錯誤: {e}")
                                
            # 🚀 再次更新 Web UI 資料檔案
            generate_interactive_main_page(current_display_account, current_view_mode)
        except: pass
        
        time.sleep(15)

# --- FastAPI Web 伺服器設定 ---
app = FastAPI()

@app.get("/")
def serve_html():
    html_path = os.path.join(base_dir, "Brawl_Tactics_Dashboard.html")
    if os.path.exists(html_path):
        return FileResponse(html_path)
    return HTMLResponse("<h1>系統初始化中，請稍後幾秒並重新整理頁面...</h1>", status_code=200)

@app.get("/dashboard_data.js")
def serve_dash_js():
    js_path = os.path.join(base_dir, "dashboard_data.js")
    if os.path.exists(js_path):
        return FileResponse(js_path)
    return HTMLResponse("window.__DYNAMIC_APP_DATA__ = {};", media_type="application/javascript")

@app.get("/ping_data.js")
def serve_ping_js():
    js_path = os.path.join(base_dir, "ping_data.js")
    if os.path.exists(js_path):
        return FileResponse(js_path)
    return HTMLResponse("window.__PING_DATA__ = {};", media_type="application/javascript")

# --- 主程式啟動 ---
if __name__ == "__main__":
    # 在背景啟動所有爬蟲引擎 (小號優先)
    threading.Thread(target=fetch_account_data, args=('小號',), daemon=True).start()
    threading.Thread(target=fetch_account_data, args=('大號',), daemon=True).start()
    threading.Thread(target=ping_worker, daemon=True).start()
    threading.Thread(target=brawlboard_keepalive_worker, daemon=True).start()
    
    print("\n" + "="*50)
    print("🚀 戰術主控台 Web 伺服器已成功啟動！")
    print("👉 請打開你的瀏覽器，輸入網址： http://localhost:8000")
    print("="*50 + "\n")
    
    # 啟動 FastAPI
    uvicorn.run(app, host="0.0.0.0", port=8000, log_level="error")
